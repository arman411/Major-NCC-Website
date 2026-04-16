"""
excel_utils.py — NCC Website
Bulk Excel Export and Import utilities using openpyxl.

Exports:
  • export_students_excel(students)   → BytesIO workbook stream
  • export_attendance_excel(records)  → BytesIO workbook stream

Import:
  • parse_bulk_students(file_stream)  → list[dict] validated rows or raises ValueError
"""

import io
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook # type: ignore
    from openpyxl.styles import ( # type: ignore
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter # type: ignore
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ─────────────────────────────────────────────────────
#  Shared helpers
# ─────────────────────────────────────────────────────
NAVY   = "0D2B5E"
SKY    = "3498DB"
WHITE  = "FFFFFF"
HEADER = "2980B9"


def _style_header_row(ws, row_num: int, num_cols: int):
    """Apply dark-blue header styling to a given row."""
    fill  = PatternFill("solid", fgColor=NAVY)
    font  = Font(color=WHITE, bold=True, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin  = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = fill
        cell.font   = font
        cell.alignment = align
        cell.border = border


def _style_data_row(ws, row_num: int, num_cols: int):
    """Alternating row colour."""
    bg = "EBF5FB" if row_num % 2 == 0 else WHITE
    fill  = PatternFill("solid", fgColor=bg)
    align = Alignment(vertical="center", wrap_text=True)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = fill
        cell.alignment = align


def _auto_width(ws):
    """Auto-fit each column to its content."""
    for col_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in col_cells) + 4
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length, 40)


# ─────────────────────────────────────────────────────
#  Student Export
# ─────────────────────────────────────────────────────
STUDENT_HEADERS = [
    "S.No", "Roll No", "Full Name", "Branch", "Year",
    "NCC Wing", "Phone", "Email", "Gender", "Status",
    "Cadet No", "Enrolled On"
]


def export_students_excel(students) -> io.BytesIO:
    """
    Convert a list of Student model objects into a formatted Excel workbook.
    Returns a BytesIO stream ready to be sent as a file download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "NCC Cadets"
    ws.row_dimensions[1].height = 30

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(STUDENT_HEADERS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"NCC Unit — Govt. Polytechnic Hamirpur (HP)  |  Cadet Roster  |  {datetime.now().strftime('%d %b %Y')}"
    title_cell.font  = Font(bold=True, size=13, color=WHITE)
    title_cell.fill  = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for col_idx, header in enumerate(STUDENT_HEADERS, start=1):
        ws.cell(row=2, column=col_idx, value=header)
    _style_header_row(ws, 2, len(STUDENT_HEADERS))
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, student in enumerate(students, start=3):
        ws.cell(row=row_idx, column=1,  value=row_idx - 2)
        ws.cell(row=row_idx, column=2,  value=student.roll_no)
        ws.cell(row=row_idx, column=3,  value=f"{student.first_name} {student.last_name}")
        ws.cell(row=row_idx, column=4,  value=student.branch)
        ws.cell(row=row_idx, column=5,  value=student.year)
        ws.cell(row=row_idx, column=6,  value=student.ncc_wing)
        ws.cell(row=row_idx, column=7,  value=student.phone)
        ws.cell(row=row_idx, column=8,  value=student.email)
        ws.cell(row=row_idx, column=9,  value=student.gender)
        ws.cell(row=row_idx, column=10, value=student.status.upper())
        ws.cell(row=row_idx, column=11, value=student.cadet_no or "—")
        ws.cell(row=row_idx, column=12, value=student.enrolled_at.strftime('%d-%m-%Y') if student.enrolled_at else "—")
        _style_data_row(ws, row_idx, len(STUDENT_HEADERS))

    # Freeze first two rows and auto-width
    ws.freeze_panes = "A3"
    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────
#  Attendance Export
# ─────────────────────────────────────────────────────
ATT_HEADERS = [
    "S.No", "Roll No", "Cadet Name", "Branch",
    "Date", "Parade Type", "Status", "Remarks"
]


def export_attendance_excel(records) -> io.BytesIO:
    """
    Convert AttendanceRecord objects into a formatted Excel sheet.
    Returns BytesIO stream.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(ATT_HEADERS))}1")
    tc = ws["A1"]
    tc.value = f"NCC GPH Hamirpur — Attendance Register  |  Exported: {datetime.now().strftime('%d %b %Y %H:%M')}"
    tc.font  = Font(bold=True, size=12, color=WHITE)
    tc.fill  = PatternFill("solid", fgColor=NAVY)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Headers
    for col_idx, h in enumerate(ATT_HEADERS, start=1):
        ws.cell(row=2, column=col_idx, value=h)
    _style_header_row(ws, 2, len(ATT_HEADERS))

    # Data
    for idx, rec in enumerate(records, start=3):
        student = rec.student
        ws.cell(row=idx, column=1, value=idx - 2)
        ws.cell(row=idx, column=2, value=student.roll_no if student else "—")
        ws.cell(row=idx, column=3, value=f"{student.first_name} {student.last_name}" if student else "—")
        ws.cell(row=idx, column=4, value=student.branch if student else "—")
        ws.cell(row=idx, column=5, value=rec.date)
        ws.cell(row=idx, column=6, value=rec.parade_type)
        ws.cell(row=idx, column=7, value="✔ Present" if rec.present else "✘ Absent")
        ws.cell(row=idx, column=8, value=rec.remarks or "")
        _style_data_row(ws, idx, len(ATT_HEADERS))

        # Colour code present/absent
        status_cell = ws.cell(row=idx, column=7)
        if rec.present:
            status_cell.font = Font(color="1E8449", bold=True)
        else:
            status_cell.font = Font(color="C0392B", bold=True)

    ws.freeze_panes = "A3"
    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────
#  Bulk Import — parse student Excel rows
# ─────────────────────────────────────────────────────
REQUIRED_IMPORT_COLS = {
    "roll_no", "first_name", "last_name", "dob",
    "gender", "phone", "email", "branch", "year", "ncc_wing"
}

IMPORT_TEMPLATE_HEADERS = [
    "roll_no", "first_name", "last_name", "dob", "gender",
    "phone", "email", "branch", "year", "ncc_wing",
    "prev_experience", "address", "motivation"
]


def generate_import_template() -> io.BytesIO:
    """Generate a blank Excel import template for admins to fill in."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cadet Import Template"

    instructions = (
        "Fill in each row with one cadet's data. "
        "Required columns: roll_no, first_name, last_name, dob (YYYY-MM-DD), "
        "gender, phone, email, branch, year, ncc_wing. "
        "Do NOT delete the header row."
    )
    ws.merge_cells(f"A1:{get_column_letter(len(IMPORT_TEMPLATE_HEADERS))}1")
    inst_cell = ws["A1"]
    inst_cell.value = instructions
    inst_cell.font  = Font(italic=True, color="7F8C8D", size=10)
    inst_cell.fill  = PatternFill("solid", fgColor="FDFEFE")
    inst_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 40

    for col_idx, h in enumerate(IMPORT_TEMPLATE_HEADERS, start=1):
        ws.cell(row=2, column=col_idx, value=h)
    _style_header_row(ws, 2, len(IMPORT_TEMPLATE_HEADERS))

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_bulk_students(file_stream) -> list:
    """
    Parse an uploaded .xlsx file stream into a list of validated dicts.
    Raises ValueError with a descriptive message on validation failure.
    """
    try:
        wb = load_workbook(file_stream, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read Excel file: {e}")

    ws = wb.active # type: ignore
    rows = list(ws.iter_rows(values_only=True)) # type: ignore

    if len(rows) < 2:
        raise ValueError("Excel file must have a header row and at least one data row.")

    # Find headers — assume they're in row 1 or 2 (template has an instructions row on row 1)
    header_row_idx = 0
    for i in range(min(3, len(rows))):
        if rows[i] and "roll_no" in [str(c).strip().lower() if c else "" for c in rows[i]]:
            header_row_idx = i
            break
    else:
        raise ValueError("Header row with 'roll_no' column not found in first 3 rows.")

    headers = [str(c).strip().lower() if c else "" for c in rows[header_row_idx]] # type: ignore

    # Validate required columns exist
    missing = REQUIRED_IMPORT_COLS - set(headers)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    parsed = []
    errors = [] # type: ignore
    for row_num, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2): # type: ignore
        if not any(row):    # Skip blank rows
            continue

        data = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))}

        # Validate required fields
        missing_vals = [col for col in REQUIRED_IMPORT_COLS if not data.get(col)]
        if missing_vals:
            errors.append(f"Row {row_num}: missing values for {', '.join(missing_vals)}")
            continue

        parsed.append({
            "roll_no":         data["roll_no"],
            "first_name":      data["first_name"],
            "last_name":       data["last_name"],
            "dob":             data["dob"],
            "gender":          data["gender"],
            "phone":           data["phone"],
            "email":           data["email"],
            "branch":          data["branch"],
            "year":            data["year"],
            "ncc_wing":        data["ncc_wing"],
            "prev_experience": data.get("prev_experience", ""),
            "address":         data.get("address", ""),
            "motivation":      data.get("motivation", ""),
        })

    if errors:
        raise ValueError(f"Validation errors in {len(errors)} rows:\n" + "\n".join(errors[:10])) # type: ignore

    if not parsed:
        raise ValueError("No valid data rows found in the uploaded file.")

    return parsed
